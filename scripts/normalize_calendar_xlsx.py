# -*- coding: utf-8 -*-
"""Gera a base normalizada do calendário (docs/calendario_base/calendario_base.json)
a partir do xlsx de trabalho da Duofy.

Uso:
    pip install openpyxl
    python scripts/normalize_calendar_xlsx.py "caminho/para/Calendario.xlsx"

Depois, para carregar no banco:
    python -m app.import_calendar   (a partir de apps/api, com DATABASE_URL apontando ao banco)

Regras de mapeamento documentadas em docs/calendario_base/README.md.
"""
from __future__ import annotations

import datetime as dt
import json
import os
import sys

import openpyxl

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs", "calendario_base")


def s(v):
    return "" if v is None else str(v).strip()


def clean_multiline(v):
    return " ".join(s(v).split())


def detect_brands(text: str):
    up = (text or "").upper()
    if "3 REDES" in up or "TODAS" in up or up.strip() == "TODOS":
        return ["duofy_solucoes", "postos_combustiveis", "deathcare"]
    order = []
    for tok, slug in [("POSTO", "postos_combustiveis"), ("DEATH", "deathcare"),
                      ("DUOFY", "duofy_solucoes")]:
        pos = up.find(tok)
        if pos >= 0:
            order.append((pos, slug))
    order.sort()
    out = []
    for _, slug in order:
        if slug not in out:
            out.append(slug)
    if not out and "TOTVS" in up:
        out = ["duofy_solucoes"]
    return out


def map_brand_primary(produto: str, perfil: str = ""):
    up = (perfil or "").upper()
    feed_brand = None
    if "(FEED)" in up:
        b = detect_brands(up.split("(FEED)")[0])
        if b:
            feed_brand = b[0]
    prod = detect_brands(produto)
    perf = detect_brands(perfil)
    allb = []
    for x in (([feed_brand] if feed_brand else []) + prod + perf):
        if x and x not in allb:
            allb.append(x)
    primary = feed_brand or (prod[0] if prod else (perf[0] if perf else "duofy_solucoes"))
    return primary, allb


def map_channel(canal: str):
    up = (canal or "").upper()
    found = []
    for tok, name in [("INSTA", "Instagram"), ("LINKEDIN", "LinkedIn"), ("LINKEIDN", "LinkedIn"),
                      ("LIKEDIN", "LinkedIn"), ("YOUTUBE", "YouTube"), ("WHATS", "WhatsApp"),
                      ("BLOG", "Blog"), ("E-MAIL", "E-mail"), ("EMAIL", "E-mail")]:
        if tok in up and name not in found:
            found.append(name)
    return " + ".join(found) if found else ("Instagram" if up else None)


def map_format(fmt: str):
    up = (fmt or "").upper()
    for tok, name in [("CARROSSEL", "Carrossel"), ("REELS", "Reels"), ("REEL", "Reels"),
                      ("MOTION", "Motion"), ("STORY", "Story"), ("CARD", "Card"),
                      ("EBOOK", "Ebook"), ("ARTIGO", "Artigo"), ("BLOG", "Artigo"),
                      ("INFOGRAF", "Infográfico"), ("FOTO", "Foto"), ("VIDEO", "Vídeo"),
                      ("VÍDEO", "Vídeo"), ("POST", "Post")]:
        if tok in up:
            return name
    return "Post" if up else None


def map_status(st: str):
    up = (st or "").upper()
    if not up:
        return "planned"
    if "DIVULGAD" in up or "PUBLICAD" in up:
        return "completed"
    if "CANCEL" in up or "FORA DO AR" in up:
        return "cancelled"
    if any(k in up for k in ["EDI", "PRODUZINDO", "PRODUZIR", "GRAVAD", "EM PRODU"]):
        return "in_progress"
    if any(k in up for k in ["AGENDAD", "PRONTO", "AGUARDANDO", "PRÓXIM", "PROXIM"]):
        return "scheduled"
    return "planned"


def parse_date(v, force_year=None):
    if isinstance(v, dt.datetime):
        d = v
    elif isinstance(v, dt.date):
        d = dt.datetime(v.year, v.month, v.day)
    else:
        return None
    if force_year and d.year < 2000:
        try:
            d = d.replace(year=force_year)
        except ValueError:
            d = d.replace(year=force_year, day=28)
    return d.isoformat()


def build(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    def sheet(sub):
        for ws in wb.worksheets:
            if sub.lower() in ws.title.lower():
                return ws
        return None

    events = []

    def add_event(sheet_name, *, date_raw, force_year, title, tema, texto, comemorativa,
                  produto, canal, formato, perfil, status):
        title = clean_multiline(title)
        if not title:
            return
        primary, all_brands = map_brand_primary(produto, perfil)
        iso = parse_date(date_raw, force_year)
        desc = [p for p in [clean_multiline(tema), clean_multiline(texto)] if p]
        payload = {
            "origem_aba": sheet_name, "produto_original": s(produto),
            "canal_original": s(canal), "formato_original": s(formato),
            "perfil": clean_multiline(perfil), "comemorativa": clean_multiline(comemorativa),
        }
        if len(all_brands) > 1:
            payload["marcas_adicionais"] = all_brands[1:]
        events.append({
            "brand_slug": primary, "category": "content", "title": title[:255],
            "description": " — ".join(desc)[:4000], "event_type": "content",
            "channel": map_channel(canal), "format": map_format(formato),
            "start_at": iso, "status": map_status(status), "date_missing": iso is None,
            "execution_payload": {k: v for k, v in payload.items() if v},
        })

    ws = sheet("2º tri")
    for r in ws.iter_rows(min_row=2, values_only=True):
        add_event("2º tri 2026", date_raw=r[1], force_year=2026, title=r[3], tema=r[4], texto=r[7],
                  comemorativa=r[2], produto=r[6], canal=r[8], formato=r[5], perfil=r[9], status=r[10])
    ws = sheet("1º tri")
    for r in ws.iter_rows(min_row=2, values_only=True):
        add_event("1º tri 2026", date_raw=r[1], force_year=2026, title=r[3], tema=r[4], texto=r[6],
                  comemorativa=r[2], produto=None, canal=r[7], formato=r[5], perfil=r[8], status=r[9])
    ws = sheet("2025")
    for r in ws.iter_rows(min_row=2, values_only=True):
        add_event("2025", date_raw=r[1], force_year=None, title=r[3], tema=r[2], texto=r[5],
                  comemorativa=None, produto=r[6], canal=r[6] or r[7], formato=r[4], perfil=r[7], status=r[8])
    ws = sheet("editorial julho")
    for r in ws.iter_rows(min_row=2, values_only=True):
        add_event("EDITORIAL JULHO", date_raw=r[1], force_year=None, title=r[2], tema=r[3], texto=r[4],
                  comemorativa=None, produto=None, canal=None, formato=r[6], perfil=None, status=None)

    themes = []
    ws = sheet("banco de temas")
    for r in ws.iter_rows(min_row=2, values_only=True):
        titulo = clean_multiline(r[2]); tema = clean_multiline(r[1])
        if not titulo and not tema:
            continue
        primary, _ = map_brand_primary(r[3])
        themes.append({
            "title": (titulo or tema)[:255], "theme": tema[:2000],
            "brand_slug": primary if s(r[3]) else None, "audience": clean_multiline(r[4]),
            "kind": clean_multiline(r[5]) or clean_multiline(r[0]),
            "owner": clean_multiline(r[6]), "status": s(r[7]),
        })

    roteiros = []
    ws = sheet("roteiros")
    for r in ws.iter_rows(min_row=3, values_only=True):
        tema = clean_multiline(r[4]); roteiro = clean_multiline(r[5])
        if not tema and not roteiro:
            continue
        primary, _ = map_brand_primary(r[2])
        roteiros.append({
            "title": (tema or "Roteiro")[:255], "brand_slug": primary if s(r[2]) else None,
            "recording_status": clean_multiline(r[3]), "script": roteiro[:8000],
            "scenes": clean_multiline(r[6]), "lettering": clean_multiline(r[7]),
            "caption": clean_multiline(r[8]), "status": s(r[9]),
        })

    return {"events": events, "themes": themes, "roteiros": roteiros}


def main():
    if len(sys.argv) < 2:
        print("uso: python scripts/normalize_calendar_xlsx.py <caminho_do_xlsx>")
        sys.exit(1)
    base = build(sys.argv[1])
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "calendario_base.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(base, f, ensure_ascii=False, indent=2)
    print(f"OK: {len(base['events'])} eventos, {len(base['themes'])} temas, "
          f"{len(base['roteiros'])} roteiros -> {out}")


if __name__ == "__main__":
    main()
